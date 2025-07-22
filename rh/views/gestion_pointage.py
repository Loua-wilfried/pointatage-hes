from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime, date, timedelta, time
import qrcode
import io
import base64
from rh.models import Pointage
from rh.forms import PointageForm, EmployeForm
from institutions.models import Employe

from institutions.models import Employe


@login_required
@permission_required('rh.delete_pointage', raise_exception=True)
def pointage_delete(request, pk):
    pointage = get_object_or_404(Pointage, pk=pk)
    if request.method == 'POST':
        pointage.delete()
        return redirect('pointages_list')
    return render(request, 'rh/pointage/pointage_confirm_delete.html', {'pointage': pointage})

@login_required
@permission_required('rh.view_pointage', raise_exception=True)
def pointages_list(request):
    from roles_permissions.models import Role
    from institutions.models import Employe
    from agences.models import Agence
    pointages = Pointage.objects.select_related('employe', 'employe__role', 'agence').all()

    # Filtres dynamiques
    employe_id = request.GET.get('employe')
    fonction_id = request.GET.get('fonction')
    agence_id = request.GET.get('agence')
    telephone = request.GET.get('telephone')
    date_selected = request.GET.get('date')
    heure_min = request.GET.get('heure_min')
    heure_max = request.GET.get('heure_max')
    type_pointage = request.GET.get('type')
    source = request.GET.get('source')
    statut = request.GET.get('statut')

    # Gestion des périodes
    import calendar
    
    periode = request.GET.get('periode', 'jour')
    date_fin_selected = request.GET.get('date_fin')
    
    if date_selected:
        date_debut = datetime.strptime(date_selected, "%Y-%m-%d").date()
    else:
        date_debut = datetime.now().date()
    # Calculer la plage de dates selon la période
    if periode == 'jour':
        date_fin = date_debut
    elif periode == 'semaine':
        # Début de semaine (lundi)
        jours_depuis_lundi = date_debut.weekday()
        date_debut = date_debut - timedelta(days=jours_depuis_lundi)
        date_fin = date_debut + timedelta(days=6)
    elif periode == 'mois':
        # Début et fin du mois
        date_debut = date_debut.replace(day=1)
        if date_debut.month == 12:
            date_fin = date_debut.replace(year=date_debut.year + 1, month=1) - timedelta(days=1)
        else:
            date_fin = date_debut.replace(month=date_debut.month + 1) - timedelta(days=1)
    elif periode == 'trimestre':
        # Début et fin du trimestre
        mois_actuel = date_debut.month
        if mois_actuel <= 3:  # Q1: Jan-Mar
            date_debut = date_debut.replace(month=1, day=1)
            date_fin = date_debut.replace(month=3, day=31)
        elif mois_actuel <= 6:  # Q2: Apr-Jun
            date_debut = date_debut.replace(month=4, day=1)
            date_fin = date_debut.replace(month=6, day=30)
        elif mois_actuel <= 9:  # Q3: Jul-Sep
            date_debut = date_debut.replace(month=7, day=1)
            date_fin = date_debut.replace(month=9, day=30)
        else:  # Q4: Oct-Dec
            date_debut = date_debut.replace(month=10, day=1)
            date_fin = date_debut.replace(month=12, day=31)
    elif periode == 'annee':
        # Début et fin de l'année
        date_debut = date_debut.replace(month=1, day=1)
        date_fin = date_debut.replace(month=12, day=31)
    elif periode == 'personnalise' and date_fin_selected:
        date_fin = datetime.strptime(date_fin_selected, "%Y-%m-%d").date()
    else:
        date_fin = date_debut
    
    # Pour compatibilité avec le code existant
    date_cible = date_debut.isoformat()

    # Appliquer les filtres employés/fonction/agence
    employes = Employe.objects.all()
    if employe_id:
        employes = employes.filter(id=employe_id)
    if fonction_id:
        employes = employes.filter(role_id=fonction_id)
    if agence_id:
        employes = employes.filter(agence_id=agence_id)
    if telephone:
        employes = employes.filter(telephone=telephone)

    # Récupérer tous les pointages du jour filtré
    pointages = Pointage.objects.select_related('employe', 'employe__role', 'agence').filter(date=date_cible)
    if type_pointage:
        pointages = pointages.filter(type=type_pointage)
    if heure_min:
        pointages = pointages.filter(heure__gte=heure_min)
    if heure_max:
        pointages = pointages.filter(heure__lte=heure_max)
    # On n'applique plus le filtre source ici, mais après la normalisation plus bas

    # Calcul du statut pour chaque employé sur la période
    from types import SimpleNamespace
    pointages_list = []
    statistiques_data = {'presences': 0, 'a_lheure': 0, 'retards': 0, 'absences': 0, 'departs_anticipes': 0, 'sorties_respectees': 0}
    
    # Pour chaque employé, analyser chaque jour de la période
    current_date = date_debut
    while current_date <= date_fin:
        for emp in employes:
            entrees = Pointage.objects.filter(employe=emp, date=current_date, type='entree').order_by('heure')
            sorties = Pointage.objects.filter(employe=emp, date=current_date, type__in=['depart', 'sortie']).order_by('heure')
            if entrees.exists():
                premier_entree = entrees.first()
                if premier_entree.heure <= time(8, 0, 0):
                    statut_entree = 'a_lheure'
                    statistiques_data['a_lheure'] += 1
                else:
                    statut_entree = 'retard'
                    statistiques_data['retards'] += 1
                premier_entree.statut = statut_entree
                pointages_list.append(premier_entree)
                statistiques_data['presences'] += 1
                
                if sorties.exists():
                    derniere_sortie = sorties.last()
                    if derniere_sortie.heure < time(17, 0, 0):
                        statut_sortie = 'depart_anticipe'
                        statistiques_data['departs_anticipes'] += 1
                    elif time(17, 0, 0) <= derniere_sortie.heure <= time(18, 0, 0):
                        statut_sortie = 'sortie_respectee'
                        statistiques_data['sorties_respectees'] += 1
                    else:
                        statut_sortie = 'depart_supplementaire'
                    derniere_sortie.statut = statut_sortie
                    pointages_list.append(derniere_sortie)
            else:
                absent_virtual = SimpleNamespace()
                absent_virtual.employe = emp
                absent_virtual.date = current_date
                absent_virtual.heure = None
                absent_virtual.type = ''
                absent_virtual.source = ''
                absent_virtual.agence = getattr(emp, 'agence', None)
                absent_virtual.statut = 'absent'
                pointages_list.append(absent_virtual)
                statistiques_data['absences'] += 1
        current_date += timedelta(days=1)

    # Mapping des variantes de source vers la valeur canonique
    def normalize_source(val):
        v = (val or '').strip().lower()
        if v in ['qr', 'qr code', 'code qr', 'qrcode', 'qr_code']:
            return 'qr_code'
        if v in ['geolocalisation', 'géolocalisation', 'geoloc', 'geo', 'geo_localisation']:
            return 'geolocalisation'
        if v in ['manuel', 'manuelle', 'saisie manuelle', 'main', 'm']:
            return 'manuel'
        return v

    if source:
        source_norm = normalize_source(source)
        pointages_list = [p for p in pointages_list if normalize_source(getattr(p, 'source', '')) == source_norm]

    # Filtre sur le statut
    if statut == 'retard':
        pointages_list = [p for p in pointages_list if p.statut == 'Retard']
    elif statut == 'absent':
        pointages_list = [p for p in pointages_list if p.statut == 'Absent']
    elif statut == 'present':
        pointages_list = [p for p in pointages_list if p.statut == 'Présent']
    elif statut == 'depart_anticipe':
        pointages_list = [p for p in pointages_list if p.statut == 'depart_anticipe']
    elif statut == 'sortie_respectee':
        pointages_list = [p for p in pointages_list if p.statut == 'sortie_respectee']

    # Tri robuste : heure None en dernier
    def tri_pointage(p):
        from datetime import time
        heure = p.heure if p.heure is not None else time(0, 0)
        return (p.date, heure)
    pointages_list = sorted(pointages_list, key=tri_pointage, reverse=True)

    # Préparer les listes pour les filtres
    employes = Employe.objects.all().order_by('nom')
    fonctions = Role.objects.all().order_by('nom_role')
    agences = Agence.objects.all().order_by('nom')

    # Export PDF/Excel (détection, à implémenter plus tard)
    export_type = request.GET.get('export')
    if export_type == 'pdf':
        # Export PDF avec reportlab
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.pdfgen import canvas
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="pointages.pdf"'
        c = canvas.Canvas(response, pagesize=landscape(letter))
        width, height = landscape(letter)
        y = height - 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30, y, "Liste des pointages filtrés")
        y -= 30
        c.setFont("Helvetica-Bold", 10)
        headers = ["Employé", "Fonction", "Agence", "Date", "Heure", "Type", "Source"]
        x_positions = [30, 170, 310, 450, 530, 600, 670]
        for i, h in enumerate(headers):
            c.drawString(x_positions[i], y, h)
        y -= 20
        c.setFont("Helvetica", 10)
        for p in pointages:
            if y < 40:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 10)
                for i, h in enumerate(headers):
                    c.drawString(x_positions[i], y, h)
                y -= 20
                c.setFont("Helvetica", 10)
            employe = str(p.employe)
            fonction = getattr(getattr(p.employe, 'role', None), 'nom_role', '')
            agence = p.agence.nom if p.agence else ''
            date = p.date.strftime('%d/%m/%Y')
            heure = p.heure.strftime('%H:%M')
            type_disp = dict(p.TYPE_CHOICES).get(p.type, p.type)
            source_disp = dict(p.SOURCE_CHOICES).get(p.source, p.source)
            values = [employe, fonction, agence, date, heure, type_disp, source_disp]
            for i, v in enumerate(values):
                c.drawString(x_positions[i], y, str(v))
            y -= 18
        c.save()
        return response
    elif export_type == 'excel':
        # Export Excel avec openpyxl
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pointages"
        headers = ["Employé", "Fonction", "Agence", "Date", "Heure", "Type", "Source"]
        ws.append(headers)
        for p in pointages:
            employe = str(p.employe)
            fonction = getattr(getattr(p.employe, 'role', None), 'nom_role', '')
            agence = p.agence.nom if p.agence else ''
            date = p.date.strftime('%d/%m/%Y')
            heure = p.heure.strftime('%H:%M')
            type_disp = dict(p.TYPE_CHOICES).get(p.type, p.type)
            source_disp = dict(p.SOURCE_CHOICES).get(p.source, p.source)
            ws.append([employe, fonction, agence, date, heure, type_disp, source_disp])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=pointages.xlsx'
        wb.save(response)
        return response

    return render(request, 'rh/pointage/pointages_list.html', {
        'pointages': pointages_list,
        'employes': employes,
        'fonctions': fonctions,
        'agences': agences,
        'statistiques': statistiques_data,
        'periode_info': {
            'periode': periode,
            'date_debut': date_debut,
            'date_fin': date_fin,
        }
    })

@login_required
@permission_required('rh.add_pointage', raise_exception=True)
def pointage_create(request):
    if request.method == 'POST':
        form = PointageForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('pointages_list')
    else:
        form = PointageForm()
    return render(request, 'rh/pointage/pointage_form.html', {'form': form, 'action': 'Ajouter'})

@login_required
@permission_required('rh.change_pointage', raise_exception=True)
def pointage_update(request, pk):
    pointage = get_object_or_404(Pointage, pk=pk)
    if request.method == 'POST':
        form = PointageForm(request.POST, instance=pointage)
        if form.is_valid():
            form.save()
            return redirect('pointages_list')
    else:
        form = PointageForm(instance=pointage)
    return render(request, 'rh/pointage/pointage_form.html', {'form': form, 'action': 'Modifier'})

# --- AJOUT EMPLOYÉ ---
from django.contrib import messages

@login_required
@permission_required('institutions.add_employe', raise_exception=True)
def ajouter_employe(request):
    if request.method == 'POST':
        form = EmployeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Employé ajouté avec succès.")
            return redirect('pointages_list')
    else:
        form = EmployeForm()
    return render(request, 'rh/pointage/ajouter_employer.html', {'form': form})

# --- VUE TABLEAU DE BORD POINTAGE ---
from django import forms
from django.utils import timezone
from types import SimpleNamespace
from django.contrib.auth.decorators import login_required, permission_required

class TableauBordPointageFilterForm(forms.Form):
    employe = forms.ChoiceField(label="Employé", required=False)
    date = forms.DateField(label="Date", required=False, widget=forms.DateInput(attrs={'type': 'date'}))
    agence = forms.ChoiceField(label="Agence", required=False)
    fonction = forms.ChoiceField(label="Fonction", required=False)
    statut = forms.ChoiceField(label="Statut", required=False, choices=[
        ("", "Tous"),
        ("present", "Présent"),
        ("retard", "Retard"),
        ("absent", "Absent"),
        ("depart_anticipe", "Départ anticipé"),
        ("sortie_respectee", "Sortie respectée"),
    ])

    def __init__(self, *args, **kwargs):
        employes = kwargs.pop('employes')
        agences = kwargs.pop('agences')
        fonctions = kwargs.pop('fonctions')
        super().__init__(*args, **kwargs)
        self.fields['employe'].choices = [('', 'Tous')] + [(str(e.id), str(e)) for e in employes]
        self.fields['agence'].choices = [('', 'Toutes')] + [(str(a.id), a.nom) for a in agences]
        self.fields['fonction'].choices = [('', 'Toutes')] + [(str(f.id), f.nom_role) for f in fonctions]

@login_required
@permission_required('rh.view_pointage', raise_exception=True)
def tableau_de_bord_pointage(request):
    from roles_permissions.models import Role
    from institutions.models import Employe
    from agences.models import Agence
    from rh.models import Pointage
    from datetime import date as date_cls, datetime
    from types import SimpleNamespace

    # Préparation des listes pour les filtres
    employes = Employe.objects.all().order_by('nom')
    fonctions = Role.objects.all().order_by('nom_role')
    agences = Agence.objects.all().order_by('nom')

    # Récupération des paramètres GET
    GET = request.GET.copy()
    form = TableauBordPointageFilterForm(GET or None, employes=employes, agences=agences, fonctions=fonctions)

    employe_id = GET.get('employe')
    fonction_id = GET.get('fonction')
    agence_id = GET.get('agence')
    telephone = GET.get('telephone')
    date_selected = GET.get('date')
    heure_min = GET.get('heure_min')
    heure_max = GET.get('heure_max')
    type_pointage = GET.get('type')
    source = GET.get('source')
    statut = GET.get('statut')

    # Date cible (par défaut aujourd'hui)
    from django.utils import timezone
    if date_selected:
        try:
            date_cible = datetime.strptime(date_selected, "%Y-%m-%d").date()
        except Exception:
            date_cible = timezone.localdate()
    else:
        date_cible = timezone.localdate()

    # Filtres employés
    filtered_employes = employes
    if employe_id:
        filtered_employes = filtered_employes.filter(id=employe_id)
    if fonction_id:
        filtered_employes = filtered_employes.filter(role_id=fonction_id)
    if agence_id:
        filtered_employes = filtered_employes.filter(agence_id=agence_id)
    if telephone:
        filtered_employes = filtered_employes.filter(telephone=telephone)

    # Récupérer tous les pointages du jour filtré
    pointages = Pointage.objects.select_related('employe', 'employe__role', 'agence').filter(date=date_cible)
    if type_pointage:
        pointages = pointages.filter(type=type_pointage)
    if heure_min:
        pointages = pointages.filter(heure__gte=heure_min)
    if heure_max:
        pointages = pointages.filter(heure__lte=heure_max)

    # Calcul du statut pour chaque pointage réel
    pointages_list = []
    employe_ids_pointes = set()
    for p in pointages:
        statut_val = 'present'
        if p.type == 'entree' and p.heure > p.heure.__class__.fromisoformat('08:00:00'):
            statut_val = 'retard'
        elif not Pointage.objects.filter(employe=p.employe, date=p.date, type='entree').exists():
            statut_val = 'absent'
        elif p.type in ['depart', 'sortie']:
            if p.heure < p.heure.__class__.fromisoformat('17:00:00'):
                statut_val = 'depart_anticipe'
            else:
                statut_val = 'sortie_respectee'
        p.statut = statut_val
        pointages_list.append(p)
        employe_ids_pointes.add(p.employe.id)

    # Calcul du nombre de présents (employés ayant scanné une entrée par QR code)
    nb_presents = Pointage.objects.filter(
        date=date_cible,
        type__in=['arrivee', 'entree'],
        source__icontains='qr'
    ).values('employe').distinct().count()

    # Ajouter les absents virtuels (employés sans pointage ce jour-là)
    for emp in filtered_employes:
        if emp.id not in employe_ids_pointes:
            absent_virtual = SimpleNamespace()
            absent_virtual.employe = emp
            absent_virtual.date = date_cible
            absent_virtual.heure = None
            absent_virtual.type = ''
            absent_virtual.source = ''
            absent_virtual.agence = getattr(emp, 'agence', None)
            absent_virtual.statut = 'absent'
            pointages_list.append(absent_virtual)

    # Mapping des variantes de source vers la valeur canonique
    def normalize_source(val):
        v = (val or '').strip().lower()
        if v in ['qr', 'qr code', 'code qr', 'qrcode', 'qr_code']:
            return 'qr_code'
        if v in ['geolocalisation', 'géolocalisation', 'geoloc', 'geo', 'geo_localisation']:
            return 'geolocalisation'
        if v in ['manuel', 'manuelle', 'saisie manuelle', 'main', 'm']:
            return 'manuel'
        return v

    if source:
        source_norm = normalize_source(source)
        pointages_list = [p for p in pointages_list if normalize_source(getattr(p, 'source', '')) == source_norm]

    # Filtre sur le statut
    if statut:
        pointages_list = [p for p in pointages_list if p.statut == statut]

    # Statistiques temps réel
    # 'present' = tous ceux qui ont pointé entrée dans la journée (heure indifférente)
    nb_present = Pointage.objects.filter(date=date_cible, type='entree').values('employe').distinct().count()
    # 'ontime' = présents dont le premier pointage entrée est <= 08:00:00
    emps_ontime = set()
    for emp in filtered_employes:
        entrees = Pointage.objects.filter(employe=emp, date=date_cible, type='entree').order_by('heure')
        if entrees.exists() and entrees.first().heure <= entrees.first().heure.__class__.fromisoformat('08:00:00'):
            emps_ontime.add(emp.id)
    stats = {
        'present': nb_present,
        'ontime': len(emps_ontime),
        'retard': sum(1 for p in pointages_list if p.statut == 'retard'),
        'absent': sum(1 for p in pointages_list if p.statut == 'absent'),
        'depart_anticipe': sum(1 for p in pointages_list if p.statut == 'depart_anticipe'),
        'sortie_respectee': sum(1 for p in pointages_list if p.statut == 'sortie_respectee'),
    }
    # Ajout des taux pour le dashboard
    nb_employes = filtered_employes.count() if hasattr(filtered_employes, 'count') else len(filtered_employes)
    nb_present = stats['present']
    nb_retard = stats['retard']
    nb_absent = stats['absent']
    stats['taux_presence'] = round((nb_present / nb_employes) * 100, 2) if nb_employes else 0
    stats['taux_retard'] = round((nb_retard / nb_employes) * 100, 2) if nb_employes else 0
    stats['absents_sans_justif'] = nb_absent  # À adapter si besoin

    # Tri par nom employé
    pointages_list = sorted(pointages_list, key=lambda p: (str(p.employe)))

    # Calcul des alertes absents/retards pour la journée sélectionnée
    total = stats['present'] + stats['absent'] + stats['retard']
    alerte_absents = total > 0 and stats['absent'] / total > 0.1
    alerte_retards = total > 0 and stats['retard'] / total > 0.3

    # --- Données réelles pour les graphiques et l'analyse prédictive RH ---
    from datetime import timedelta
    import json
    # 7 derniers jours
    labels = [(date_cible - timedelta(days=i)).strftime('%d/%m') for i in range(6, -1, -1)]
    # Graphiques : taux de présence/retard réels
    presence_rates = []
    retard_rates = []
    depart_anticipe_rates = []
    for i in range(6, -1, -1):
        d = date_cible - timedelta(days=i)
        pts = Pointage.objects.filter(date=d)
        nb_emps = filtered_employes.count() if hasattr(filtered_employes, 'count') else len(filtered_employes)
        nb_present = pts.filter(type='entree', heure__lte='08:00:00').count()
        nb_retard = pts.filter(type='entree', heure__gt='08:00:00').count()
        nb_depart_anticipe = pts.filter(type__in=['depart', 'sortie'], heure__lt='17:00:00').count()
        presence_rates.append(round((nb_present / nb_emps)*100, 2) if nb_emps else 0)
        retard_rates.append(round((nb_retard / nb_emps)*100, 2) if nb_emps else 0)
        depart_anticipe_rates.append(round((nb_depart_anticipe / nb_emps)*100, 2) if nb_emps else 0)
    # Analyse prédictive RH réelle
    employes_risque = []
    for emp in filtered_employes:
        nb_retards = 0
        nb_absences = 0
        for i in range(7):
            d = date_cible - timedelta(days=i)
            pt = Pointage.objects.filter(employe=emp, date=d, type='entree').first()
            if pt:
                if pt.heure > pt.heure.__class__.fromisoformat('08:00:00'):
                    nb_retards += 1
            else:
                nb_absences += 1
        score = max(0, 100 - 10*nb_retards - 20*nb_absences)
        employes_risque.append({
            "nom": str(emp),
            "retards": nb_retards,
            "score": score
        })
    employes_risque = sorted(employes_risque, key=lambda x: x['score'])[:3]
    # Niveau de risque global
    niveau = "Faible"
    couleur_risque = "#22c55e"
    if any(e['score'] < 60 for e in employes_risque):
        niveau = "Élevé"
        couleur_risque = "#ef4444"
    elif any(e['score'] < 75 for e in employes_risque):
        niveau = "Modéré"
        couleur_risque = "#f59e42"
    resume_pred = f"{niveau} risque d’absence détecté : {', '.join(e['nom'] for e in employes_risque)}."

    # --- Variables avancées pour enrichir le dashboard ---
    # Prédiction IA fictive
    proba_absence = 15  # À remplacer par une vraie prédiction IA si dispo

    # Top 3 ponctuels (meilleur score, déjà trié dans employes_risque)
    top_ponctuels = sorted(employes_risque, key=lambda x: -x['score'])[:3]

    # Top 3 absents sur 7 jours
    absences_7j = {}
    for emp in filtered_employes:
        nb_abs = 0
        for i in range(7):
            d = date_cible - timedelta(days=i)
            pt = Pointage.objects.filter(employe=emp, date=d, type='entree').first()
            if not pt:
                nb_abs += 1
        absences_7j[emp] = nb_abs
    top_absents = sorted(absences_7j.items(), key=lambda x: -x[1])[:3]

    # Top 3 retardataires sur 7 jours
    retards_7j = {}
    for emp in filtered_employes:
        nb_ret = 0
        for i in range(7):
            d = date_cible - timedelta(days=i)
            pt = Pointage.objects.filter(employe=emp, date=d, type='entree').first()
            if pt and pt.heure > pt.heure.__class__.fromisoformat('08:00:00'):
                nb_ret += 1
        retards_7j[emp] = nb_ret
    top_retardataires = sorted(retards_7j.items(), key=lambda x: -x[1])[:3]

    # Top 3 départs anticipés sur 7 jours
    dep_7j = {}
    for emp in filtered_employes:
        nb_dep = 0
        for i in range(7):
            d = date_cible - timedelta(days=i)
            pt = Pointage.objects.filter(employe=emp, date=d, type__in=['depart', 'sortie']).order_by('heure').first()
            if pt and pt.heure < pt.heure.__class__.fromisoformat('17:00:00'):
                nb_dep += 1
        dep_7j[emp] = nb_dep
    top_depart_anticipe = sorted(dep_7j.items(), key=lambda x: -x[1])[:3]

    # Top 3 présents sur 7 jours
    pres_7j = {}
    for emp in filtered_employes:
        nb_pre = 0
        for i in range(7):
            d = date_cible - timedelta(days=i)
            pt = Pointage.objects.filter(employe=emp, date=d, type='entree').first()
            if pt and pt.heure <= pt.heure.__class__.fromisoformat('08:00:00'):
                nb_pre += 1
        pres_7j[emp] = nb_pre
    top_present = sorted(pres_7j.items(), key=lambda x: -x[1])[:3]

    # Stat GPS activé (fictif)
    stats['gps_activé'] = True
    # Premier pointage du jour
    pts_today = Pointage.objects.filter(date=date_cible).order_by('heure')
    stats['premier_pointage'] = pts_today.first().heure.strftime('%H:%M') if pts_today.exists() else None
    # Dernier pointage du jour
    stats['dernier_pointage'] = pts_today.last().heure.strftime('%H:%M') if pts_today.exists() else None

    # --- Préparation des données pour le graphe comparaison par agence ---
    comparaison_agences_labels = []
    comparaison_agences_data = []
    agences_all = agences.distinct()
    for agence in agences_all:
        employes_agence = filtered_employes.filter(agence=agence)
        nb_emps = employes_agence.count()
        nb_present = Pointage.objects.filter(date=date_cible, employe__in=employes_agence, type='entree', heure__lte='08:00:00').count() if nb_emps else 0
        taux = round((nb_present / nb_emps) * 100, 2) if nb_emps else 0
        comparaison_agences_labels.append(agence.nom)
        comparaison_agences_data.append(taux)

    return render(request, 'rh/pointage/tableau_de_bord_pointage.html', {
        'pointages': pointages_list,
        'employes': employes,
        'fonctions': fonctions, 
        'agences': agences,
        'stats': stats,
        'request': request,
        'chart_labels': json.dumps(labels),
        'chart_presence': json.dumps(presence_rates),
        'chart_retard': json.dumps(retard_rates),
        'chart_depart_anticipe': json.dumps(depart_anticipe_rates),
        'niveau_risque': niveau,
        'couleur_risque': couleur_risque,
        'employes_risque': employes_risque,
        'resume_pred': resume_pred,
        'alerte_absents': alerte_absents,
        'alerte_retards': alerte_retards,
        'proba_absence': proba_absence,
        'top_ponctuels': top_ponctuels,
        'top_absents': top_absents,
        'top_retardataires': top_retardataires,
        'top_depart_anticipe': top_depart_anticipe,
        'top_present': top_present,
        'comparaison_agences_labels': comparaison_agences_labels,
        'comparaison_agences_data': comparaison_agences_data,
        'nb_presents': nb_presents,
    })


from django.views.decorators.clickjacking import xframe_options_exempt

from agences.models import Agence
from django import forms

class QRCodeAgenceForm(forms.Form):
    agence = forms.ModelChoiceField(queryset=Agence.objects.filter(statut='active'), label="Agence")
    type = forms.ChoiceField(choices=[('entree', 'Entrée'), ('sortie', 'Sortie')], label="Type de pointage")

@login_required
@permission_required('rh.view_pointage', raise_exception=True)
@xframe_options_exempt
def generate_qr_code(request):
    qr_code_url = None
    agence_nom = None
    type_pointage = None
    agence_id = None
    form = None
    try:
        if request.method == 'POST':
            form = QRCodeAgenceForm(request.POST)
            if form.is_valid():
                agence = form.cleaned_data['agence']
                type_pointage = form.cleaned_data['type']
                agence_id = agence.id
                agence_nom = agence.nom
                # On encode dans le QR un JSON avec agence_id, agence_nom, type
                import json
                qr_payload = json.dumps({
                    "agence_id": agence_id,
                    "agence_nom": agence_nom,
                    "type": type_pointage
                })
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(qr_payload)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                buffer = io.BytesIO()
                img.save(buffer, format="PNG")
                qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
                qr_code_url = f"data:image/png;base64,{qr_code_base64}"
        else:
            form = QRCodeAgenceForm()
    except Exception as e:
        import logging
        logging.error(f"Erreur dans generate_qr_code : {e}")
        if not form:
            form = QRCodeAgenceForm()
    template = 'generate_qr_code.html'
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('iframe') == '1':
        template = 'qr_code_iframe.html'
    return render(request, template, {
        'form': form,
        'qr_code_url': qr_code_url,
        'agence_nom': agence_nom,
        'type_pointage': type_pointage,
        'agence_id': agence_id
    })
